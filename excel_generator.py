#!/usr/bin/env python3
"""
Excel export module for Duplicate Patient Manager

This module generates Excel reports for duplicate patient processing results.
"""

import logging
import os
from datetime import datetime
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """Generate Excel reports for duplicate patient processing"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def generate_excel_report(self, results: Dict[str, Any], patients_data: List[Dict[str, Any]], output_dir: str = None) -> str:
        """Generate Excel report from processing results and patient data"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"duplicate_patient_report_{timestamp}.xlsx"
            
            # Use output directory if provided, otherwise current directory
            if output_dir:
                filepath = os.path.join(output_dir, filename)
            else:
                filepath = filename
            
            workbook = Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create sheets
            self._create_summary_sheet(workbook, results)
            self._create_duplicate_pairs_sheet(workbook, patients_data, results)
            self._create_patients_sheet(workbook, patients_data, results)
            self._create_processing_details_sheet(workbook, results)
            
            # Save workbook
            workbook.save(filepath)
            
            self.logger.info(f"Excel report generated: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}")
            return ""
    
    def _create_summary_sheet(self, workbook: Workbook, results: Dict[str, Any]) -> None:
        """Create summary sheet"""
        sheet = workbook.create_sheet("Summary", 0)
        
        # Headers
        headers = [
            ("Duplicate Patient Management Report", "A1"),
            ("Generated On", "A3"),
            ("PG Company ID", "A4"),
            ("Total Patients", "A5"),
            ("Duplicate Groups Found", "A6"),
            ("Patients Processed", "A7"),
            ("Patients Deleted", "A8"),
            ("Orders Moved", "A9"),
            ("CC Notes Moved", "A10"),
            ("Errors Count", "A11")
        ]
        
        values = [
            ("", "B1"),
            (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "B3"),
            (results.get('pg_company_id', ''), "B4"),
            (results.get('total_patients', 0), "B5"),
            (results.get('duplicate_groups_found', 0), "B6"),
            (results.get('patients_processed', 0), "B7"),
            (results.get('patients_deleted', 0), "B8"),
            (results.get('orders_moved', 0), "B9"),
            (results.get('cc_notes_moved', 0), "B10"),
            (len(results.get('errors', [])), "B11")
        ]
        
        # Apply headers
        for header, cell in headers:
            sheet[cell] = header
            sheet[cell].font = Font(bold=True)
        
        # Apply values
        for value, cell in values:
            sheet[cell] = value
        
        # Title formatting
        sheet["A1"].font = Font(size=16, bold=True)
        sheet.merge_cells("A1:B1")
        
        # Auto-size columns
        self._auto_size_columns(sheet)
    
    def _create_duplicate_pairs_sheet(self, workbook: Workbook, patients_data: List[Dict[str, Any]], results: Dict[str, Any]) -> None:
        """Create duplicate pairs sheet showing kept and deleted patients side by side"""
        sheet = workbook.create_sheet("Duplicate Pairs")
        
        # Headers for kept and deleted patients side by side
        headers = [
            "Group #",
            # Kept Patient columns
            "KEPT - Patient ID", "KEPT - Full Name", "KEPT - DOB", "KEPT - MRN", 
            "KEPT - Total Orders", "KEPT - Status", "KEPT - Start of Care",
            # Deleted Patient columns  
            "DELETED - Patient ID", "DELETED - Full Name", "DELETED - DOB", "DELETED - MRN",
            "DELETED - Total Orders", "DELETED - Status", "DELETED - Start of Care",
            # Summary columns
            "Orders Moved", "CC Notes Moved", "PDF Verification"
        ]
        
        # Add headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            
            # Color code kept vs deleted columns
            if "KEPT" in header:
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                cell.font = Font(color="006400", bold=True)  # Dark green
            elif "DELETED" in header:
                cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
                cell.font = Font(color="8B0000", bold=True)  # Dark red
            else:
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Blue
                cell.font = Font(color="FFFFFF", bold=True)  # White
        
        # Create patient lookup dictionary
        patients_dict = {p.get('id'): p for p in patients_data}
        
        # Add duplicate pair data
        row_num = 2
        for group_num, detail in enumerate(results.get('processing_details', []), 1):
            primary_patient_id = detail.get('primary_patient_id')
            deleted_patient_ids = detail.get('deleted_patient_ids', [])
            
            primary_patient = patients_dict.get(primary_patient_id, {})
            primary_agency = primary_patient.get('agencyInfo', {})
            
            # For each deleted patient, create a row showing the pair
            for deleted_id in deleted_patient_ids:
                deleted_patient = patients_dict.get(deleted_id, {})
                deleted_agency = deleted_patient.get('agencyInfo', {})
                
                # Get PDF verification info if available
                pdf_verification = ""
                if detail.get('pdf_extractions'):
                    for extraction in detail.get('pdf_extractions', []):
                        if extraction.get('patient_id') == primary_patient_id:
                            pdf_verification = f"Primary Score: {extraction.get('score', 0)}"
                        elif extraction.get('patient_id') == deleted_id:
                            pdf_verification += f" | Deleted Score: {extraction.get('score', 0)}"
                
                row_data = [
                    group_num,
                    # Kept patient data
                    primary_patient_id,
                    f"{primary_agency.get('patientFName', '')} {primary_agency.get('patientLName', '')}".strip(),
                    primary_agency.get('dob', ''),
                    primary_agency.get('medicalRecordNo', ''),
                    primary_agency.get('totalOrders', ''),
                    primary_agency.get('patientStatus', ''),
                    primary_agency.get('startOfCare', ''),
                    # Deleted patient data
                    deleted_id,
                    f"{deleted_agency.get('patientFName', '')} {deleted_agency.get('patientLName', '')}".strip(),
                    deleted_agency.get('dob', ''),
                    deleted_agency.get('medicalRecordNo', ''),
                    deleted_agency.get('totalOrders', ''),
                    deleted_agency.get('patientStatus', ''),
                    deleted_agency.get('startOfCare', ''),
                    # Summary data
                    detail.get('moved_orders', 0),
                    detail.get('moved_cc_notes', 0),
                    pdf_verification.strip(" |")
                ]
                
                # Add row data with formatting
                for col_num, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    
                    # Apply row-level formatting
                    if col_num >= 2 and col_num <= 8:  # Kept patient columns
                        cell.fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")  # Very light green
                    elif col_num >= 9 and col_num <= 15:  # Deleted patient columns
                        cell.fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")  # Very light red
                
                row_num += 1
        
        # Auto-size columns
        self._auto_size_columns(sheet)
        
        # Add borders
        self._add_borders(sheet, row_num - 1, len(headers))
    
    def _create_patients_sheet(self, workbook: Workbook, patients_data: List[Dict[str, Any]], results: Dict[str, Any]) -> None:
        """Create patients data sheet"""
        sheet = workbook.create_sheet("Patients Data")
        
        # Headers
        headers = [
            "Patient ID", "Full Name", "First Name", "Last Name", "DOB", "MRN", 
            "Company ID", "PG Company ID", "Total Orders", "Status", "Action Taken",
            "Start of Care", "Service Line", "State", "Payor Source"
        ]
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add patient data
        row_num = 2
        deleted_patient_ids = set()
        kept_patient_ids = set()
        
        # Get deleted and kept patient IDs from processing details
        for detail in results.get('processing_details', []):
            deleted_patient_ids.update(detail.get('deleted_patient_ids', []))
            kept_patient_ids.add(detail.get('primary_patient_id'))
        
        for patient in patients_data:
            agency_info = patient.get('agencyInfo', {})
            patient_id = patient.get('id', '')
            
            # Determine action taken
            if patient_id in deleted_patient_ids:
                action_taken = "DELETED"
                status_color = "FFCCCB"  # Light red
            elif patient_id in kept_patient_ids:
                action_taken = "KEPT (Primary)"
                status_color = "90EE90"  # Light green
            else:
                action_taken = "NO ACTION"
                status_color = "FFFFFF"  # White
            
            row_data = [
                patient_id,
                f"{agency_info.get('patientFName', '')} {agency_info.get('patientLName', '')}".strip(),
                agency_info.get('patientFName', ''),
                agency_info.get('patientLName', ''),
                agency_info.get('dob', ''),
                agency_info.get('medicalRecordNo', ''),
                agency_info.get('companyId', ''),
                agency_info.get('pgcompanyID', ''),
                agency_info.get('totalOrders', ''),
                agency_info.get('patientStatus', ''),
                action_taken,
                agency_info.get('startOfCare', ''),
                agency_info.get('serviceLine', ''),
                agency_info.get('state', ''),
                agency_info.get('payorSource', '')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                
                # Color code based on action
                if col_num == 11:  # Action Taken column
                    cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                    if action_taken == "DELETED":
                        cell.font = Font(color="8B0000", bold=True)  # Dark red
                    elif action_taken == "KEPT (Primary)":
                        cell.font = Font(color="006400", bold=True)  # Dark green
            
            row_num += 1
        
        # Auto-size columns
        self._auto_size_columns(sheet)
        
        # Add borders
        self._add_borders(sheet, row_num - 1, len(headers))
    
    def _create_processing_details_sheet(self, workbook: Workbook, results: Dict[str, Any]) -> None:
        """Create processing details sheet"""
        sheet = workbook.create_sheet("Processing Details")
        
        # Headers
        headers = ["Group #", "Primary Patient ID", "Primary Patient Name", "Deleted Patient IDs", 
                  "Orders Moved", "CC Notes Moved", "Errors"]
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add processing details
        row_num = 2
        for detail in results.get('processing_details', []):
            row_data = [
                detail.get('group_number', ''),
                detail.get('primary_patient_id', ''),
                detail.get('primary_patient_name', ''),
                ', '.join(detail.get('deleted_patient_ids', [])),
                detail.get('moved_orders', 0),
                detail.get('moved_cc_notes', 0),
                '; '.join(detail.get('errors', [])) if detail.get('errors') else 'None'
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
            
            row_num += 1
        
        # Auto-size columns
        self._auto_size_columns(sheet)
        
        # Add borders
        self._add_borders(sheet, row_num - 1, len(headers))
        
        # Add errors sheet if there are errors
        if results.get('errors'):
            self._create_errors_sheet(workbook, results.get('errors', []))
    
    def _create_errors_sheet(self, workbook: Workbook, errors: List[str]) -> None:
        """Create errors sheet"""
        sheet = workbook.create_sheet("Errors")
        
        # Header
        sheet["A1"] = "Error Messages"
        sheet["A1"].font = Font(bold=True)
        sheet["A1"].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        sheet["A1"].font = Font(color="FFFFFF", bold=True)
        
        # Add errors
        for i, error in enumerate(errors, 2):
            sheet[f"A{i}"] = error
        
        # Auto-size columns
        self._auto_size_columns(sheet)
    
    def _auto_size_columns(self, sheet) -> None:
        """Auto-size columns based on content"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_borders(self, sheet, max_row: int, max_col: int) -> None:
        """Add borders to table"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                sheet.cell(row=row, column=col).border = thin_border
